

# =============================================================================
# Import modules
# =============================================================================
# Import general purpose module(s)
import pandas as pd
import openpyxl as op
# Import module-specific module/file(s)
import nicexcel.utils as utils
import nicexcel.const as const


# =============================================================================
# filter worksheet content
# =============================================================================
def filter_sheet(ws: op.worksheet.worksheet.Worksheet,
                 header_row: int,
                 first_col: int,
                 last_col: int):
    # get header rows as string
    hr_str = utils.get_row_as_str(row=header_row)
    # get letter of first column
    first_col_str = op.utils.get_column_letter(first_col)
    # max column letter
    last_col_str = op.utils.get_column_letter(last_col)
    #
    ws.auto_filter.ref = first_col_str + hr_str + ':' + last_col_str + \
                         utils.get_row_as_str(ws.max_row)
    # return back worksheet
    return ws


# =============================================================================
# freeze header
# =============================================================================
def freeze_header(ws: op.worksheet.worksheet.Worksheet,
                  header_row: int,
                  first_col: int):
    # get header rows as string
    hr_p1_str = utils.get_row_as_str(row=header_row+1)
    # get first column as string
    first_col_str = op.utils.get_column_letter(first_col)
    # give concatenation to freeze_panes function
    ws.freeze_panes = 'A' + hr_p1_str
    # return back worksheet
    return ws


# =============================================================================
# auto format width of all columns
# =============================================================================
def set_column_width(df: pd.core.frame.DataFrame,
                     ws: op.worksheet.worksheet.Worksheet,
                     first_col: int,
                     first_data_col: int,
                     last_col: int):

    # loop through all columns
    for idx in range(first_col, last_col+1):
        # for index columns
        if idx in range(first_col, first_data_col):
            # if multi index case
            if isinstance(df.index, pd.MultiIndex):
                series = pd.Series(df.index.levels[idx-first_col])
            # if single index case
            else:
                series = df.index
        # for data columns
        else:
            series = df[df.columns[idx - first_data_col]]
        # len of largest item or column header
        max_len = max((
            series.astype(str).map(len).max(),  # len of largest item
            len(str(series.name))  # len of column name/header
        )) + const.READABILITY_MARGIN  # adding a little extra space
        # if larger than maximum allowed value, cap it
        if max_len >= const.MAX_COLUMN_WIDTH:
            max_len = const.MAX_COLUMN_WIDTH
        # get column letter equivalent of number
        idx_letter = op.utils.get_column_letter(idx)
        # set column width
        ws.column_dimensions[idx_letter].width = max_len

    return ws


# =============================================================================
# apply number format to each column as required
# =============================================================================
def apply_format_to_column(ws: op.worksheet.worksheet.Worksheet,
                           field: str,
                           header_dict: dict,
                           number_format: str,
                           header_row: int):
    # loop over rows of the worksheet
    for row in range(header_row, ws.max_row + 1):
        # apply specified number format
        ws.cell(row=row, column=header_dict[field]).number_format \
            = number_format
    # return back worksheet object
    return ws


#
def apply_dtypes_formats(df: pd.core.frame.DataFrame,
                         ws: op.worksheet.worksheet.Worksheet,
                         header_row: int,
                         header_dict: dict):
    # get data types of dataframe
    dtypes_ser = df.dtypes
    # iterate over columns
    for field in dtypes_ser.index:
        # if integer
        if 'int' in dtypes_ser[field].name:
            ws = apply_format_to_column(ws=ws,
                                        header_dict=header_dict,
                                        field=field,
                                        number_format=const.OP_INTEGER,
                                        header_row=header_row)
        elif 'float' in dtypes_ser[field].name:
            ws = apply_format_to_column(ws=ws,
                                        header_dict=header_dict,
                                        field=field,
                                        number_format=const.OP_NORMAL,
                                        header_row=header_row)
    #
    return ws


# add description
def apply_user_formats(ws: op.worksheet.worksheet.Worksheet,
                       cols_format: dict,
                       header_row: int,
                       header_dict: dict):

    # format columns
    for form in cols_format.keys():
        if form in const.NF_PERCENT:
            # scan over rows, method here
            for field in cols_format[form]:
                # set format to each column
                ws = apply_format_to_column(ws=ws, header_dict=header_dict,
                                            field=field,
                                            number_format=const.OP_PERCENT,
                                            header_row=header_row)
        # float with commas (2 decimals) with thousand separator
        elif form in const.NF_NORMAL:
            # scan over rows, method here
            for field in cols_format[form]:
                # set format to each column
                ws = apply_format_to_column(ws=ws, header_dict=header_dict,
                                            field=field,
                                            number_format=const.OP_NORMAL,
                                            header_row=header_row)
        # integer with no commas with thousand separator
        elif form in const.NF_INTEGER:
            # scan over rows, method here
            for field in cols_format[form]:
                # set format to each column
                ws = apply_format_to_column(ws=ws, header_dict=header_dict,
                                            field=field,
                                            number_format=const.OP_INTEGER,
                                            header_row=header_row)
    #
    return ws


# =============================================================================
# auto format width of all columns
# =============================================================================
def format_columns(df: pd.core.frame.DataFrame,
                   ws: op.worksheet.worksheet.Worksheet,
                   dtypes_format: bool,
                   cols_format: dict,
                   header_row: int):
    # read header as dictionary, per each column name get corresponding
    # column number
    header_dict = utils.read_header(ws=ws,
                                    max_col=ws.max_column,
                                    first_row=header_row)
    # apply pandas dtypes formatting
    if dtypes_format:
        ws = apply_dtypes_formats(df=df,
                                  ws=ws,
                                  header_row=header_row,
                                  header_dict=header_dict)
    # apply user specified formatting
    ws = apply_user_formats(ws=ws,
                            cols_format=cols_format,
                            header_row=header_row,
                            header_dict=header_dict)

    # return back worksheet
    return ws
