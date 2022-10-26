"""
Module description, consistent with __init__ file
=====

Provides

  - Module contents:
      Module that takes Pandas DataFrame objects and writes them to excel
        in a nicely formatted files:
         - column width auto-adapted to fit characters contained in it
        - filterable Excel columns
        - header freezed by default
        - no indexing by default
        - number formatting of columns according to dataframe data types
        - possibility of specifying number formatting

How to use the documentation
----------------------------
Documentation is available in docstrings provided with the code

Required external packages
--------
`openpyxl`, `pandas`
"""


# =============================================================================
# Import modules
# =============================================================================
# Import general purpose module(s)
import pandas as pd
import openpyxl as op
import os
# Import module-specific module/file(s)
import nicexcel.excel_format as xf


# =============================================================================
# Modules check
# =============================================================================
# Verify that all modules are present
# _ = module.module_check(dependencies=__dependencies__,
#                         ifmissing=_MODULE_MISSING,
#                         ifversion=_MODULE_VERSION)


# =============================================================================
# get worksheet coordinates
# =============================================================================
def get_worksheet_coordinates(df, index, *args, **kwargs):
    # get starts of first row and last column
    header_row, first_col = get_worksheet_start(index=index, *args, **kwargs)
    # get first data column (excluding index columns)
    first_data_col = get_worksheet_first_data_col(df=df, first_col=first_col,
                                                  index=index)
    # get last column
    last_col = get_worksheet_end(df=df, first_data_col=first_data_col,
                                 index=index)
    # return back all coordinates parameters
    return header_row, first_col, first_data_col, last_col


# =============================================================================
# get worksheet start
# =============================================================================
def get_worksheet_start(index, *args, **kwargs):
    if 'startrow' in kwargs:
        header_row = kwargs['startrow'] + 1
    else:
        header_row = 1

    # get initial column both as number and as character
    if 'startcol' in kwargs:
        import openpyxl as op
        # check if index is in keywords arguments
        first_col = kwargs['startcol'] + 1
    else:
        first_col = 1

    return header_row, first_col


# =============================================================================
# get index of first data column
# =============================================================================
#
def get_worksheet_first_data_col(df, first_col, index):
    if index:
        # get number of columns of dataframe index
        len_index = df.index.nlevels
        # index is boolean but still can sum integers and binaries (ok)
        first_data_col = first_col + len_index
    else:
        # if no index all columns are data columns
        first_data_col = first_col
    # last column of worksheet
    return first_data_col


# =============================================================================
# get worksheet end column
# =============================================================================
# get last column of worksheet
def get_worksheet_end(df, first_data_col, index):
    # get shape of dataframe
    num_cols = len(df.columns)
    # index is boolean but still can sum integers and binaries (ok)
    last_col = first_data_col + num_cols - 1
    # last column of worksheet
    return last_col


# =============================================================================
# orchestrator of function formatting
# =============================================================================
def format_sheet(df: pd.core.frame.DataFrame,
                 filename: str,
                 sheet_name: str,
                 filter_on: bool,
                 header_freezed: bool,
                 header_row: int,
                 first_col: int,
                 first_data_col: int,
                 last_col: int,
                 dtypes_format: bool,
                 cols_format: dict):

    # get openpyxl representation of dataframe
    wb = op.load_workbook(filename=filename)
    # get worksheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # activate filter
    if filter_on:
        ws = xf.filter_sheet(ws=ws, header_row=header_row,
                             first_col=first_col,
                             last_col=last_col)

    # number format of columns
    ws = xf.format_columns(df=df,
                           ws=ws,
                           dtypes_format=dtypes_format,
                           cols_format=cols_format,
                           header_row=header_row)

    # set column width format with max width
    ws = xf.set_column_width(df=df, ws=ws, first_col=first_col,
                             first_data_col=first_data_col,
                             last_col=last_col)

    # freeze header ( optional )
    if header_freezed:
        ws = xf.freeze_header(ws=ws, header_row=header_row,
                              first_col=first_col)

    # finally save back workbook
    wb.save(filename=filename)


# =============================================================================
# function for writing one dataframe on excel (equivalent of pd.to_excel(...) )
# =============================================================================
def to_excel(df: pd.core.frame.DataFrame,
             filename: str,
             sheet_name: str = None,
             index: bool = False,
             filter_on: bool = True,
             header_freezed: bool = True,
             dtypes_format: bool = True,
             cols_format: dict = {},
             *args,
             **kwargs):
    """
    Nicely export a Pandas DataFrame to an Excel file.

    The function takes as inputs the pandas dataframes and the file path
    where the file will be generated. Then it is possible to specify a set
    of parameters that will condition the output formatting of the excel file

    Parameters
    ----------
    df : pd.core.frame.DataFrame
        Pandas DataFrame instance to be written. It is suggested that the
        dataframe size is inferior to MS Excel limits (on .xlsx files
        1,048,576 rows by 16,384 columns)
    filename : str
        String containing full file path of the Excel file of output; .xlsx
        extension required, if missing, the program automatically corrects
        it back to .xlsx
    sheet_name: str
        Optional parameter containing name of the sheet where the Dataframe
        will be written on
    index: bool
        Optional binary parameter specifying whether index will be exported
    filter_on: bool
        Optional binary parameter specifying whether excel file should be
        filterable or not
    header_freezed: Optional binary parameter specifying whether excel file
        should have header row freezed or not
    dtypes_format: Optional binary parameter specifying whether
        filw with number formats corresponding to Pandas DataFrame
        column data types
    cols_format: dict
        Optional dictionary specifying per each output data type,
        the columns that should be formatted accordingly. If
        "dtypes_format" is True, then this dictionary
        over-writes Pandas specifications


    Returns
    -------

    """
    # checks on arguments
    filename_, file_extension = os.path.splitext(filename)
    if file_extension != '.xlsx':
        print('Warning, you passed a file path without .xlsx extension, '
              'exporting file to ' + filename_ + '.xlsx')
        filename = filename_ + '.xlsx'

    if sheet_name is None:
        df.to_excel(excel_writer=filename, engine='openpyxl', index=index,
                    *args, **kwargs)
    else:
        df.to_excel(excel_writer=filename, engine='openpyxl',
                    index=index, sheet_name=sheet_name, *args, **kwargs)

    # get coordinates of dataframe, row of header and columns for data
    header_row, first_col, first_data_col, last_col = \
        get_worksheet_coordinates(df=df, index=index, *args, **kwargs)

    # format sheet using openpyxl, test this
    format_sheet(df=df,
                 filename=filename,
                 sheet_name=sheet_name,
                 filter_on=filter_on,
                 header_freezed=header_freezed,
                 header_row=header_row,
                 first_col=first_col,
                 first_data_col=first_data_col,
                 last_col=last_col,
                 dtypes_format=dtypes_format,
                 cols_format=cols_format)
    # terminate function
    return None


# =============================================================================
# function for writing multiple dataframes on different sheets of same excel
# =============================================================================
def to_excel_ms(dfs: dict,
                filename: str,
                index: bool = False,
                filter_on: bool = True,
                header_freezed: bool = True,
                dtypes_format: bool = True,
                dict_cols_format: dict = {},
                *args,
                **kwargs):
    """
    Nicely export a dictionary of dataframes to an Excel file on multiple
    sheets.

    Extended description of function...

    Parameters
    ----------
    dfs : dict {str: pd.core.frame.DataFrame}
        Dictionary where each key is a string and each value is a Pandas
        DataFrame instance. The method will export each DataFrame on a
        separate worksheet of the same Excel output file, where the sheets
        name will equal the key string of the DataFrame in dfs. It is
        suggested that the DataDrame sizes are inferior to MS Excel limits (
        on .xlsx files 1,048,576 rows by 16,384 columns)
    filename : str
        String containing full file path of the Excel file of output; .xlsx
        extension required, if missing, the program automatically corrects
        it back to .xlsx
    index: bool
        Optional binary parameter specifying whether index will be exported
    filter_on: bool
        Optional binary parameter specifying whether excel file should be
        filterable or not
    header_freezed: Optional binary parameter specifying whether excel file
        should have header row freezed or not
    dtypes_format: Optional binary parameter specifying whether
        filw with number formats corresponding to Pandas DataFrame
        column data types
    dict_cols_format: dict {str: dict}
        Optional dictionary having sheetnames as keys and dictionaries as
        values. These dictionaries contain as keys output data types and as
        values the columns that should be formatted accordingly. If
        "dtypes_format" is True, then this dictionary
        over-writes Pandas specifications

    Returns
    -------

    """
    # init workbook with openpyxl engine
    excel_writer = pd.ExcelWriter(filename, engine='openpyxl')

    # write each worksheet separately with pd.ExcelWriter
    for sheet_name in dfs.keys():
        #
        df = dfs[sheet_name]
        #
        df.to_excel(excel_writer=excel_writer, index=index,
                    sheet_name=sheet_name , *args, **kwargs)
    # save excel
    excel_writer.save()

    # format each worksheet separately
    for sheet_name, df in dfs.items():
        print(sheet_name)
        # get coordinates of dataframe, row of header and columns for data
        header_row, first_col, first_data_col, last_col = \
            get_worksheet_coordinates(df=df, index=index)
        # , *args, **kwargs)

        # check if contained this sheet_name
        if sheet_name not in dict_cols_format.keys():
            dict_cols_format[sheet_name] = {}

        # format sheet using openpyxl
        format_sheet(df=df,
                     filename=filename,
                     sheet_name=sheet_name,
                     filter_on=filter_on,
                     header_freezed=header_freezed,
                     header_row=header_row,
                     first_col=first_col,
                     first_data_col=first_data_col,
                     last_col=last_col,
                     dtypes_format=dtypes_format,
                     cols_format=dict_cols_format[sheet_name])
    # terminate function
    return None
